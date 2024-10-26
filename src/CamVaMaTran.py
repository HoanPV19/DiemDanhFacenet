import tensorflow as tf
from imutils.video import VideoStream
import argparse
import facenet
import imutils
import os
import pickle
import align.detect_face
import numpy as np
import cv2
import pandas as pd
import tkinter as tk
import threading  # Thêm threading để chạy song song
from openpyxl import load_workbook
SESSION_COLUMN = "B6"


def update_attendance(excel_file, msv):
    # Tải workbook và worksheet
    wb = load_workbook(excel_file)
    ws = wb.active  # Hoặc chỉ định sheet cụ thể nếu cần

    # Tìm chỉ số cột từ tên nhãn SESSION_COLUMN
    session_column_index = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == SESSION_COLUMN:
            session_column_index = col
            break

    if session_column_index is None:
        print(f"Không tìm thấy cột '{SESSION_COLUMN}' trong worksheet.")
        return

    # Kiểm tra xem cột session_column có tồn tại
    found = False
    
    for row in range(2, ws.max_row + 1):  # Giả sử hàng đầu tiên là tiêu đề
        if ws.cell(row=row, column=3).value == msv:  # Cột MSV là cột thứ 3 (C)
            ws.cell(row=row, column=session_column_index).value = 'A'  # Đánh dấu "A" vào cột session_column
            found = True

    if found:
        # Lưu workbook
        wb.save(excel_file)
        print("Attendance updated successfully.")
    else:
        print("No matching MSV found.")


# Hàm tạo cửa sổ ma trận tên
def create_matrix_window(root, names):
    matrix_frame = tk.Frame(root)
    matrix_frame.pack()

    num_names = len(names)
    
    # Chia ma trận thành 12 hàng và 8 cột 
    num_cols = 8
    num_rows = 12

    labels = {}

    for i in range(num_rows * num_cols):
        row = i // num_cols
        col = i % num_cols

        if i < num_names:
            name = names[i]
        else:
            name = ""  # Ô trống nếu không có đủ người

        # Tạo label cho mỗi người với màu nền trắng và chữ đen ban đầu
        label = tk.Label(matrix_frame, text=name, width=12, height=2, borderwidth=2, relief="solid", padx=5, pady=5, bg="white", fg="black")
        label.grid(row=row, column=col, padx=5, pady=5)
        labels[name] = label
    return labels


# Hàm cập nhật màu nền của nhãn khi một người được xác thực
def remove_name_from_matrix(labels, name):
    # Đổi màu nhãn thành màu xanh và giữ chữ đen khi đã xác thực
    if name in labels:
        labels[name].config(bg="green", fg="white")  # Đổi nền thành màu xanh và chữ vẫn màu đen


# Hàm nhận diện và cập nhật danh sách
def face_recognition_process(sess, model, class_names, images_placeholder, embeddings, phase_train_placeholder, pnet, rnet, onet, cap, labels, EXCEL_FILE_PATH):
    while True:
        frame = cap.read()
        frame = imutils.resize(frame, width=850)
        frame = cv2.flip(frame, 1)

        bounding_boxes, _ = align.detect_face.detect_face(frame, 20, pnet, rnet, onet, [0.6, 0.7, 0.7], 0.709)

        faces_found = bounding_boxes.shape[0]
        try:
            if faces_found > 0:
                det = bounding_boxes[:, 0:4]
                bb = np.zeros((faces_found, 4), dtype=np.int32)
                for i in range(faces_found):
                    bb[i][0] = det[i][0]
                    bb[i][1] = det[i][1]
                    bb[i][2] = det[i][2]
                    bb[i][3] = det[i][3]

                    if (bb[i][3] - bb[i][1]) / frame.shape[0] > 0.25:
                        cropped = frame[bb[i][1]:bb[i][3], bb[i][0]:bb[i][2],:]
                        scaled = cv2.resize(cropped, (160, 160), interpolation=cv2.INTER_CUBIC)
                        scaled = facenet.prewhiten(scaled)
                        scaled_reshape = scaled.reshape(-1, 160, 160, 3)
                        feed_dict = {images_placeholder: scaled_reshape, phase_train_placeholder: False}
                        emb_array = sess.run(embeddings, feed_dict=feed_dict)

                        predictions = model.predict_proba(emb_array)
                        best_class_indices = np.argmax(predictions, axis=1)
                        best_class_probabilities = predictions[np.arange(len(best_class_indices)), best_class_indices]
                        best_name = class_names[best_class_indices[0]]

                        # Hiển thị tên và khung tương ứng với xác suất
                        if best_class_probabilities[0] > 0.71:
                            # Tỷ lệ trên 70%, khung màu xanh
                            cv2.rectangle(frame, (bb[i][0], bb[i][1]), (bb[i][2], bb[i][3]), (0, 255, 0), 2)
                            cv2.putText(frame, best_name, (bb[i][0], bb[i][3] + 20), cv2.FONT_HERSHEY_COMPLEX_SMALL, 1, (0, 255, 0), thickness=1, lineType=2)
                            
                            # Cập nhật Excel và ma trận
                            update_attendance(EXCEL_FILE_PATH, best_name)
                            remove_name_from_matrix(labels, best_name)  # Xác thực và đổi màu nhãn trong ma trận
                        else:
                            # Tỷ lệ dưới 70%, khung màu đỏ
                            cv2.rectangle(frame, (bb[i][0], bb[i][1]), (bb[i][2], bb[i][3]), (0, 0, 255), 2)
                            cv2.putText(frame, best_name, (bb[i][0], bb[i][3] + 20), cv2.FONT_HERSHEY_COMPLEX_SMALL, 1, (0, 0, 255), thickness=1, lineType=2)

                        # Hiển thị xác suất
                        cv2.putText(frame, str(round(best_class_probabilities[0], 3)), (bb[i][0], bb[i][3] + 40), cv2.FONT_HERSHEY_COMPLEX_SMALL, 1, (0, 255, 0), thickness=1, lineType=2)

        except Exception as e:
            print(f"Error: {e}")

        cv2.imshow('Face Recognition', frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--path', help='Path of the video you want to test on.', default=0)
    parser.add_argument('--classifier', help='Path to the classifier file (.pkl) for face recognition.',
                        default='Models/vector/CN6.pkl')
    parser.add_argument('--excel_path', help='Path to the Excel file for attendance data.',
                        default='Dataset/DSLOP/CN5.xlsx')
    parser.add_argument('--session_column', help='Name of the session column in the Excel file.', required=True)
    args = parser.parse_args()

    # Sử dụng đường dẫn file mã hóa và Excel từ đối số hoặc đường dẫn mặc định
    CLASSIFIER_PATH = args.classifier
    EXCEL_FILE_PATH = args.excel_path
    global SESSION_COLUMN
    SESSION_COLUMN = args.session_column
    FACENET_MODEL_PATH = 'Models/20180402-114759.pb'
    
    # Kiểm tra sự tồn tại của file mã hóa
    if not os.path.exists(CLASSIFIER_PATH):
        print(f"Classifier file not found: {CLASSIFIER_PATH}")
        return

    # Kiểm tra sự tồn tại của file Excel
    if not os.path.exists(EXCEL_FILE_PATH):
        print(f"Excel file not found: {EXCEL_FILE_PATH}")
        return

    print(f"Using classifier: {CLASSIFIER_PATH}")
    print(f"Using Excel file: {EXCEL_FILE_PATH}")
    print(f"Session column: {SESSION_COLUMN}")

    # Tiếp tục triển khai code cho nhận diện khuôn mặt và xử lý dữ liệu điểm danh từ Excel
    # Tại đây, bạn có thể sử dụng EXCEL_FILE_PATH và SESSION_COLUMN để đọc và xử lý dữ liệu điểm danh từ cột mong muốn

    # Load The Custom Classifier
    with open(CLASSIFIER_PATH, 'rb') as file:
        model, class_names = pickle.load(file)
    print("Custom Classifier, Successfully loaded")

    tf.compat.v1.disable_eager_execution()

    with tf.Graph().as_default():
        gpu_options = tf.compat.v1.GPUOptions(per_process_gpu_memory_fraction=0.6)
        sess = tf.compat.v1.Session(config=tf.compat.v1.ConfigProto(gpu_options=gpu_options, log_device_placement=False))

        with sess.as_default():
            facenet.load_model(FACENET_MODEL_PATH)

            images_placeholder = tf.compat.v1.get_default_graph().get_tensor_by_name("input:0")
            embeddings = tf.compat.v1.get_default_graph().get_tensor_by_name("embeddings:0")
            phase_train_placeholder = tf.compat.v1.get_default_graph().get_tensor_by_name("phase_train:0")
            pnet, rnet, onet = align.detect_face.create_mtcnn(sess, "src/align")

            # Thiết lập cửa sổ Tkinter chứa ma trận tên
            root = tk.Tk()
            root.title("Attendance Matrix")
            labels = create_matrix_window(root, class_names)  # Tạo ma trận chứa tên người

            cap = VideoStream(src=0).start()

            # Tạo luồng nhận diện khuôn mặt
            recognition_thread = threading.Thread(target=face_recognition_process, args=(sess, model, class_names, images_placeholder, embeddings, phase_train_placeholder, pnet, rnet, onet, cap, labels, EXCEL_FILE_PATH))
            recognition_thread.daemon = True
            recognition_thread.start()

            # Bắt đầu giao diện Tkinter
            root.mainloop()


main()
