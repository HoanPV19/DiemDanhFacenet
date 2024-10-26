import tkinter as tk
from tkinter import ttk, simpledialog
import subprocess
import os
import shutil  # Thêm thư viện shutil để xóa thư mục không rỗng
from tkinter import messagebox
import pandas as pd  # Đảm bảo pandas đã được cài đặt
from openpyxl import load_workbook

# Đường dẫn đến thư mục chứa các file mã hóa và mã điểm danh
vector_path = 'models/vector'
base_path = 'src'
dslop_path = 'Dataset/DSlop'
facedata_path = 'Dataset/Facedata'

# To store the subprocess instance
attendance_process = None


def add_folder():
    root.config(cursor="wait")
    script_path = os.path.join(base_path, 'ThemFolderAnh.py')
    
    if os.path.exists(script_path):
        # Chạy script và chờ cho đến khi hoàn thành
        subprocess.run(['python', script_path], check=True)
        update_class_list()  # Gọi hàm update_class_list ngay sau khi script hoàn thành
    else:
        print(f"File {script_path} does not exist.")
    
    root.config(cursor="")


# Hàm thêm DSSV
def add_dssv():
    root.config(cursor="wait")
    script_path = os.path.join(base_path, 'ThemDSLop.py')
    if os.path.exists(script_path):
        subprocess.Popen(['python', script_path])
    else:
        print(f"File {script_path} does not exist.")
    root.config(cursor="")


# Hàm cập nhật danh sách các nhãn từ cột D trở đi cho các buổi học
def update_sessions(selected_class):
    root.config(cursor="wait")
    session_combobox.set('')
    session_combobox['values'] = []
    excel_file_path = os.path.join(dslop_path, f"{selected_class}.xlsx")

    if os.path.exists(excel_file_path):
        try:
            # Đọc file Excel và lấy các nhãn từ cột D trở đi
            df = pd.read_excel(excel_file_path, header=0)
            headers = df.columns[3:]  # Bắt đầu từ cột D (cột thứ tư)

            # Lọc các nhãn cột không trống và tạo danh sách lựa chọn cho combobox
            session_labels = [col for col in headers if pd.notna(col)]
            session_labels.append("Tạo buổi học mới")  # Thêm lựa chọn "Tạo buổi học mới" vào cuối danh sách

            session_combobox['values'] = session_labels
        except Exception as e:
            messagebox.showerror("Error", f"Không thể đọc file Excel: {e}")
    else:
        messagebox.showwarning("Warning", f"File Excel không tồn tại: {excel_file_path}")

    root.config(cursor="")


# Cập nhật danh sách lớp từ các file .pkl
def update_class_list():
    root.config(cursor="wait")
    files = [f for f in os.listdir(vector_path) if f.endswith('.pkl')]
    class_combobox['values'] = files
    if files:
        class_combobox.current(0)
        update_sessions(files[0].split(".")[0])
    root.config(cursor="")


# Hàm gọi khi người dùng chọn lớp trong Combobox
def on_class_selected(event):
    selected_class = class_combobox.get().split(".")[0]
    update_sessions(selected_class)

# Hàm xử lý khi chọn "Tạo buổi học mới"


def create_new_column(selected_class):
    new_column_name = simpledialog.askstring("Tạo buổi học mới", "Nhập tên buổi học mới:")
    if new_column_name:
        excel_file_path = os.path.join(dslop_path, f"{selected_class}.xlsx")
        try:
            # Sử dụng openpyxl để mở file Excel
            workbook = load_workbook(excel_file_path)
            sheet = workbook.active
            
            # Lấy tất cả các giá trị trong dòng đầu tiên (tiêu đề cột)
            header_values = [cell.value for cell in sheet[1]]  # Duyệt qua các ô trong dòng đầu tiên
            
            # Kiểm tra xem tên cột mới đã tồn tại chưa
            if new_column_name in header_values:
                messagebox.showwarning("Warning", f"Buổi học '{new_column_name}' đã tồn tại trong file {selected_class}.xlsx.")
                return
            
            # Tìm ô trống đầu tiên trong dòng tiêu đề
            empty_col = None
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=1, column=col).value is None:
                    empty_col = col
                    break
            
            # Nếu không tìm thấy ô trống, thêm vào cột tiếp theo
            if empty_col is None:
                empty_col = sheet.max_column + 1
            
            sheet.cell(row=1, column=empty_col, value=new_column_name)  # Thêm tiêu đề cột
            
            workbook.save(excel_file_path)  # Lưu lại file Excel
            
            update_sessions(selected_class)  # Cập nhật lại danh sách buổi
            messagebox.showinfo("Thành công", f"Đã thêm buổi học '{new_column_name}' vào file {selected_class}.xlsx.")
        except Exception as e:
            messagebox.showerror("Error", f"Không thể tạo buổi học mới: {e}")

            
# Hàm gọi khi người dùng chọn buổi trong Combobox
def on_session_selected(event):
    selected_session = session_combobox.get()
    if selected_session == "Tạo buổi học mới":
        selected_class = class_combobox.get().split(".")[0]
        create_new_column(selected_class)


def start_attendance():
    global attendance_process
    selected_file = class_combobox.get()
    selected_session = session_combobox.get()
    
    # Kiểm tra nếu đã chọn lớp và buổi học
    if selected_file and selected_session:
        # Kiểm tra nếu người dùng chọn "Tạo buổi học mới"
        if selected_session == "Tạo buổi học mới":
            selected_class = selected_file.split('.')[0]  # Lấy tên lớp từ selected_file
            create_new_column(selected_class)  # Gọi hàm để tạo buổi học mới
            return
        
        # Tiếp tục kiểm tra sự tồn tại của các file
        script_path = os.path.join(base_path, 'CamVaMaTran.py')
        vector_file_path = os.path.join(vector_path, selected_file)
        excel_file_path = os.path.join(dslop_path, f"{selected_file.split('.')[0]}.xlsx")
        
        # Kiểm tra sự tồn tại của các file
        if os.path.exists(script_path) and os.path.exists(vector_file_path) and os.path.exists(excel_file_path):
            try:
                # Đọc file Excel
                df = pd.read_excel(excel_file_path, engine='openpyxl')
                
                # Kiểm tra xem cột buổi học có tồn tại không
                if selected_session in df.columns:
                    session_column_name = selected_session  # Lưu tên cột
                    print(f"Session column '{session_column_name}' found.")
                else:
                    messagebox.showerror("Error", f"Buổi học {selected_session} không tồn tại trong file Excel.")
                    return
                
                root.config(cursor="wait")
                
                # Gọi CamVaMaTran.py với đối số --classifier, --excel_path và --session_column (truyền tên cột)
                attendance_process = subprocess.Popen([ 
                    'python', script_path,
                    '--classifier', vector_file_path,
                    '--excel_path', excel_file_path,
                    '--session_column', session_column_name  # Truyền tên cột
                ])
                print(f"Started attendance with classifier: {vector_file_path} and session column: {session_column_name}")
                
            except Exception as e:
                messagebox.showerror("Error", f"Không thể bắt đầu điểm danh: {e}")
            finally:
                root.config(cursor="")
        else:
            messagebox.showerror("Error", f"File {script_path}, {vector_file_path}, hoặc {excel_file_path} không tồn tại.")
    else:
        messagebox.showwarning("Warning", "Vui lòng chọn lớp và buổi học.")


# Kết thúc điểm danh
def end_attendance():
    global attendance_process
    if attendance_process:
        attendance_process.terminate()
        attendance_process = None
        print("Đã kết thúc điểm danh.")

        # Hiện cửa sổ thông báo
        messagebox.showinfo("Thông báo", "Điểm danh đã kết thúc.")
    else:
        messagebox.showinfo("Info", "Chưa có buổi điểm danh nào đang diễn ra.")


def open_excel_file():
    selected_class = class_combobox.get().split('.')[0]
    excel_file_path = os.path.join(dslop_path, f"{selected_class}.xlsx")
    
    if os.path.exists(excel_file_path):
        if os.name == 'nt':  # Kiểm tra nếu hệ điều hành là Windows
            os.startfile(excel_file_path)  # Mở file Excel
        else:
            messagebox.showwarning("Warning", "Chức năng này chỉ hỗ trợ trên Windows.")
    else:
        messagebox.showwarning("Warning", f"File Excel không tồn tại: {excel_file_path}")


def delete_class():
    # Tạo cửa sổ mới để chọn lớp cần xóa
    delete_window = tk.Toplevel(root)
    delete_window.title("Xóa lớp")
    delete_window.geometry("300x400")
    
    label = tk.Label(delete_window, text="Chọn lớp để xóa:")
    label.pack(pady=10)
    
    class_list = [f.split('.')[0] for f in os.listdir(vector_path) if f.endswith('.pkl')]
    class_listbox = tk.Listbox(delete_window)
    class_listbox.pack(padx=10, pady=10, fill="both", expand=True)
    
    for item in class_list:
        class_listbox.insert(tk.END, item)
    
    def confirm_delete():
        selected_class = class_listbox.get(tk.ACTIVE)
        if selected_class:
            # Hiển thị hộp thoại xác nhận
            confirm = messagebox.askyesno("Xác nhận", f"Bạn có chắc muốn xóa lớp '{selected_class}' không?")
            if confirm:
                vector_file_path = os.path.join(vector_path, f"{selected_class}.pkl")
                excel_file_path = os.path.join(dslop_path, f"{selected_class}.xlsx")
                facedata_folder_path = os.path.join(facedata_path, selected_class)

                # Xóa file vector
                if os.path.exists(vector_file_path):
                    os.remove(vector_file_path)
                    print(f"Đã xóa file: {vector_file_path}")
                else:
                    print(f"File không tồn tại: {vector_file_path}")

                # Xóa file Excel
                if os.path.exists(excel_file_path):
                    os.remove(excel_file_path)
                    print(f"Đã xóa file: {excel_file_path}")
                else:
                    print(f"File không tồn tại: {excel_file_path}")

                # Xóa thư mục facedata
                if os.path.exists(facedata_folder_path):
                    shutil.rmtree(facedata_folder_path)  # Sử dụng shutil.rmtree để xóa thư mục không rỗng
                    print(f"Đã xóa thư mục: {facedata_folder_path}")
                else:
                    print(f"Thư mục không tồn tại: {facedata_folder_path}")

                update_class_list()
                messagebox.showinfo("Thành công", f"Đã xóa lớp '{selected_class}'.")
                delete_window.destroy()

    confirm_button = ttk.Button(delete_window, text="Xác nhận", command=confirm_delete)
    confirm_button.pack(pady=10)


# Tạo cửa sổ chính
root = tk.Tk()
root.title("Điểm danh bằng nhận diện khuôn mặt")
root.geometry("550x250")
root.configure(bg="#f0f0f0")

frame_main = tk.Frame(root, bg="#f0f0f0")
frame_main.pack(padx=10, pady=10, fill="both", expand=True)

title_label = tk.Label(frame_main, text="Điểm danh bằng nhận diện khuôn mặt", font=("Arial", 16), bg="#f0f0f0", fg="#333")
title_label.pack(pady=5)

# Khung chứa các nút chức năng
frame_left = tk.Frame(frame_main, bg="#f0f0f0", borderwidth=2, relief="groove")
frame_left.pack(side="left", padx=10, pady=10, fill="y")

btn_folder = ttk.Button(frame_left, text="Thêm Folder ảnh", command=add_folder)
btn_folder.pack(padx=10, pady=5, fill="x")

btn_dssv = ttk.Button(frame_left, text="Thêm DSSV", command=add_dssv)
btn_dssv.pack(padx=10, pady=5, fill="x")

btn_delete_class = ttk.Button(frame_left, text="Xóa lớp", command=delete_class)
btn_delete_class.pack(padx=10, pady=5, fill="x")

# Khung chứa Combobox chọn lớp
frame_center = tk.Frame(frame_main, bg="#f0f0f0", borderwidth=2, relief="groove")
frame_center.pack(side="left", padx=10, pady=10, fill="y")

class_label = tk.Label(frame_center, text="Danh sách lớp", font=("Arial", 12), bg="#f0f0f0")
class_label.pack(pady=5)

class_combobox = ttk.Combobox(frame_center, state="readonly", font=("Arial", 10))
class_combobox.pack(padx=10, pady=5, fill="x")
class_combobox.bind("<<ComboboxSelected>>", on_class_selected)

# Combobox chọn buổi
session_label = tk.Label(frame_center, text="Chọn Buổi", font=("Arial", 12), bg="#f0f0f0")
session_label.pack(pady=5)

session_combobox = ttk.Combobox(frame_center, state="readonly", font=("Arial", 10))
session_combobox.pack(padx=10, pady=5, fill="x")
session_combobox.bind("<<ComboboxSelected>>", on_session_selected)

# Khung chứa các nút điểm danh
frame_right = tk.Frame(frame_main, bg="#f0f0f0", borderwidth=2, relief="groove")
frame_right.pack(side="left", padx=10, pady=10, fill="y")

btn_start_attendance = ttk.Button(frame_right, text="Bắt đầu điểm danh", command=start_attendance)
btn_start_attendance.pack(padx=10, pady=5, fill="x")

btn_end_attendance = ttk.Button(frame_right, text="Kết thúc điểm danh", command=end_attendance)
btn_end_attendance.pack(padx=10, pady=5, fill="x")

btn_open_excel = ttk.Button(frame_right, text="Mở Danh Sách Lớp", command=open_excel_file)
btn_open_excel.pack(pady=5)

# Cập nhật danh sách lớp và khởi động ứng dụng
update_class_list()
root.mainloop()
